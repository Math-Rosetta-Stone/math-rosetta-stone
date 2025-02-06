import mysql.connector
from mysql.connector import errorcode
from openpyxl import load_workbook
from dotenv import load_dotenv
import os
import json
import re

# Get the root directory (one level up from the script's directory)
root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Load environment variables from the .env.local file in the root directory
load_dotenv(dotenv_path=os.path.join(root_dir, '.env.local'))

print(f"DB_HOST: {os.getenv('DB_HOST')}")
print(f"DB_USER: {os.getenv('DB_USER')}")
print(f"DB_PASSWORD: {os.getenv('DB_PASSWORD')}")
print(f"DB_NAME: {os.getenv('DB_NAME')}")

# Configs for MySQL connection
config = {
    'user': os.getenv('DB_USER'), 
    'password': os.getenv('DB_PASSWORD'),  # Replace with your MySQL password
    'host': os.getenv('DB_HOST'),  # Replace with your MySQL host (e.g., 'localhost')
    'database': os.getenv('DB_NAME'),  # Replace with your MySQL database name
    'port': 3306,  # Replace with your MySQL port (default is 3306)
    'charset': 'utf8mb4',  # Provides support for a larger range of characters
    'use_unicode': True  # Prevents conversion of Unicode characters to escape sequences
}

# Connect to MySQL server
try:
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()
    print("Successfully connected to MySQL server...\n")
except mysql.connector.Error as err:
    if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
        print("Something is wrong with your username or password")
    elif err.errno == errorcode.ER_BAD_DB_ERROR:
        print("Database does not exist")
    else:
        print(err)
    exit()

# Constants for Excel columns and rows
EXCEL_FILENAME = "terms.xlsx"  # Replace with your Excel file name
SHEET_NAME = "Sheet1"  # Replace with your sheet name
ROW_HEADERS = 1  # Row containing headers
ROW_START = 2  # First row of data
ROW_END = 287  # Last row of data

# Escapes SQL's special characters from a string
def escape_special_characters(string):
    if not string:
        return string
    # Define special characters to escape (based on SQL docs)
    special_characters = r'''!@&*[]{}^:=/><-()%+?;'~|"\\'''
    # Escape each special character
    for char in special_characters:
        string = string.replace(char, f"\\{char}")
    return string

# Insert languages into the `languages` table
def insert_languages():
    languages = ["French", "Chinese", "Spanish", "Portuguese", "Hindi", "Farsi", "Marathi"]
    for lang in languages:
        insert_language_query = """
        INSERT INTO languages (name)
        VALUES (%s)
        """
        cursor.execute(insert_language_query, (lang,))
    conn.commit()
    print("Successfully inserted languages...\n")

def insert_branches():
    branches = [
        (1, 10, "Calculus"),
        (2, 8, "Arithmetic"),
        (3, 12, "Geometry"),
        (4, 6, "Trigonometry"),
        (5, 9, "Algebra"),
        (6, 5, "Logical Operators"),
        (7, 7, "Sets"),
        (8, 11, "General Math"),
        (9, 10, "Statistics and Probability"),
        (10, 8, "CS")
    ]
    for branch in branches:
        insert_branch_query = """
        INSERT INTO branch (branch_no, no_of_chapters, map_name)
        VALUES (%s, %s, %s)
        """
        cursor.execute(insert_branch_query, branch)
    conn.commit()
    print("Successfully inserted branches...\n")
    
# Insert classes into the `classes` table
def insert_classes():
    classes = [
        "MATA02", "MATA22", "MATA23", "MATA29", "MATA30",
        "MATA31", "MATA34", "MATA35", "MATA36", "MATA37",
        "MATA67", "CSCA08", "CSCA20", "CSCA48", "CSCA67"
    ]
    for class_name in classes:
        insert_class_query = """
        INSERT INTO classes (name)
        VALUES (%s)
        """
        cursor.execute(insert_class_query, (class_name,))
    conn.commit()
    print("Successfully inserted classes...\n")

def insert_excel_data_into_sql_tables():
    # Load Excel spreadsheet
    wb = load_workbook(filename=EXCEL_FILENAME, data_only=True)
    ws = wb[SHEET_NAME]

    # Find and store column numbers
    classes = [
        "MATA02", "MATA22", "MATA23", "MATA29", "MATA30",
        "MATA31", "MATA34", "MATA35", "MATA36", "MATA37",
        "MATA67", "CSCA08", "CSCA20", "CSCA48", "CSCA67"
    ]
    languages = [
        "Chinese (simple) Definition", "French Definition", "Spanish Definition",
        "Portuguese Definition", "Farsi Definition", "Hindi Definition", "Marathi Definition"
    ]
    col = 1
    class_cols = []
    language_cols = []
    term_col = None
    definition_col = None
    field_col = None
    level_col = None
    example_col = None

    while ws.cell(ROW_HEADERS, col).value is not None:
        header = ws.cell(ROW_HEADERS, col).value
        if header == "English Terms":
            term_col = col
        elif header == "English Definition":
            definition_col = col
        elif header == "Field":
            field_col = col
        elif header == "Level":
            level_col = col
        elif header == "Example":
            example_col = col
        for language in languages:
            if language in header:
                language_cols.append((col, language))
                break
        for class_name in classes:
            if class_name in header:
                class_cols.append((col, class_name))
                break
        col += 1

    # Check that all required columns are found
    if (term_col is None or definition_col is None or
            field_col is None or level_col is None or example_col is None or
            len(language_cols) == 0 or len(class_cols) == 0):
        raise Exception("Missing required columns in Excel sheet...")

    # Retrieve language IDs from the database
    cursor.execute("SELECT lang_id, name FROM languages")
    language_id_map = {name: lang_id for lang_id, name in cursor.fetchall()}

    # Retrieve class IDs from the database
    cursor.execute("SELECT class_id, name FROM classes")
    class_id_map = {name: class_id for class_id, name in cursor.fetchall()}

    # Retrieve branch IDs from the database
    cursor.execute("SELECT branch_no, map_name FROM branch")
    branch_id_map = {map_name: branch_no for branch_no, map_name in cursor.fetchall()}

    # Insert data into the `terms` table
    for row in range(ROW_START, ROW_END + 1):
        # Retrieve cell values
        term = ws.cell(row, term_col).value
        definition = ws.cell(row, definition_col).value
        field = ws.cell(row, field_col).value
        level = ws.cell(row, level_col).value
        example = ws.cell(row, example_col).value

        # Check for missing or invalid data
        if term is None or term == "0" or term == 0:
            print(f"Skipping row {row}: Missing or invalid term")
            continue
        if definition is None or definition == "0" or definition == 0:
            print(f"Skipping row {row}: Missing or invalid definition")
            continue
        if field is None or field == "0" or field == 0:
            print(f"Skipping row {row}: Missing or invalid field")
            continue
        if level is None or level == "0" or level == 0:
            print(f"Skipping row {row}: Missing or invalid level")
            continue
        if example is None or example == "0" or example == 0:
            print(f"Skipping row {row}: Missing or invalid example")
            continue

        branch_number = branch_id_map.get(field)

        # Escape special characters
        term = escape_special_characters(term)
        definition = escape_special_characters(definition)
        example = escape_special_characters(example)

        # Print term details for debugging
        print(f"Inserting row {row}: term_id={row - ROW_START + 1}, term={term}, branch_no={branch_number}, rank={level}, definition={definition}, example={example}")

        # Insert into `terms` table
        insert_term_query = """
        INSERT INTO terms (term_id, term, branch_no, `rank`, definition, example)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_term_query, (
            row - ROW_START + 1,  # term_id (assuming sequential IDs)
            term,
            branch_number,  # branch_no from the database
            level,  # rank from the Level column
            definition,
            example  # example from the Example column
        ))
        conn.commit()

        # Insert into `translations` table
        for lang_col, lang_name in language_cols:
            translation = ws.cell(row, lang_col).value
            if translation:
                translation = escape_special_characters(translation)
                # Check if the combination of term_id and lang_id already exists
                cursor.execute("""
                SELECT COUNT(*) FROM translations WHERE term_id = %s AND lang_id = %s
                """, (row - ROW_START + 1, language_id_map[lang_name.split()[0]]))
                if cursor.fetchone()[0] == 0:
                    insert_translation_query = """
                    INSERT INTO translations (term_id, lang_id, definition)
                    VALUES (%s, %s, %s)
                    """
                    cursor.execute(insert_translation_query, (
                        row - ROW_START + 1,  # term_id
                        language_id_map[lang_name.split()[0]],  # lang_id from the database
                        translation
                    ))
                    conn.commit()

        # Insert into `termToClass` table
        for class_col, class_name in class_cols:
            class_value = ws.cell(row, class_col).value
            if class_value == "Yes":
                # Check if the combination of term_id and class_id already exists
                cursor.execute("""
                SELECT COUNT(*) FROM termToClass WHERE term_id = %s AND class_id = %s
                """, (row - ROW_START + 1, class_id_map[class_name]))
                if cursor.fetchone()[0] == 0:
                    insert_term_to_class_query = """
                    INSERT INTO termToClass (term_id, class_id)
                    VALUES (%s, %s)
                    """
                    cursor.execute(insert_term_to_class_query, (
                        row - ROW_START + 1,  # term_id
                        class_id_map[class_name]  # class_id from the database
                    ))
                    conn.commit()
                    
                    
# Drop and recreate tables
def recreate_tables():
    drop_tables = [
        "DROP TABLE IF EXISTS permission",
        "DROP TABLE IF EXISTS level",
        "DROP TABLE IF EXISTS chapter",
        "DROP TABLE IF EXISTS termToClass",
        "DROP TABLE IF EXISTS translations",
        "DROP TABLE IF EXISTS terms",
        "DROP TABLE IF EXISTS languages",
        "DROP TABLE IF EXISTS classes",
        "DROP TABLE IF EXISTS branch"
    ]
    create_tables = [
        """
        CREATE TABLE branch (
            branch_no INT PRIMARY KEY,
            no_of_chapters INT NOT NULL,
            map_name VARCHAR(100) NOT NULL
        )
        """,
        """
        CREATE TABLE languages (
            lang_id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL
        )
        """,
        """
        CREATE TABLE classes (
            class_id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL
        )
        """,
        """
        CREATE TABLE terms (
            term_id INT AUTO_INCREMENT PRIMARY KEY,
            term VARCHAR(255) NOT NULL,
            branch_no INT,
            `rank` INT,
            definition TEXT,
            example TEXT
        )
        """,
        """
        CREATE TABLE translations (
            term_id INT,
            lang_id INT,
            definition TEXT,
            FOREIGN KEY (term_id) REFERENCES terms(term_id),
            FOREIGN KEY (lang_id) REFERENCES languages(lang_id),
            UNIQUE (term_id, lang_id)
        )
        """,
        """
        CREATE TABLE termToClass (
            term_id INT,
            class_id INT,
            FOREIGN KEY (term_id) REFERENCES terms(term_id),
            FOREIGN KEY (class_id) REFERENCES classes(class_id),
            UNIQUE (term_id, class_id)
        )
        """,
        """
        CREATE TABLE chapter (
            chapter_no INT,
            branch_no INT,
            no_of_minigames INT NOT NULL,
            PRIMARY KEY (chapter_no, branch_no),
            FOREIGN KEY (branch_no) REFERENCES branch(branch_no)
        )
        """,
        """
        CREATE TABLE level (
            level_no INT,
            chapter_no INT,
            branch_no INT,
            minigame_name VARCHAR(8) NOT NULL CHECK (minigame_name IN ('hangman', 'matching', 'mcq', 'logo', 'fib', 'listen', 'random')),
            x INT NOT NULL,
            y INT NOT NULL,
            PRIMARY KEY (level_no, chapter_no, branch_no),
            FOREIGN KEY (chapter_no, branch_no) REFERENCES chapter(chapter_no, branch_no),
            UNIQUE (chapter_no, branch_no, x, y),
            UNIQUE (branch_no, chapter_no, level_no)
        )
        """,
        """
        CREATE TABLE permission (
            user_id VARCHAR(255) NOT NULL,
            curr_branch_no INT NOT NULL,
            curr_chapter_no INT NOT NULL,
            curr_level_no INT NOT NULL,
            PRIMARY KEY (user_id, curr_branch_no),
            FOREIGN KEY (curr_branch_no, curr_chapter_no, curr_level_no) REFERENCES level(branch_no, chapter_no, level_no)
        )
        """
    ]

    # Execute drop and create statements
    for query in drop_tables + create_tables:
        cursor.execute(query)
    conn.commit()
    print("Successfully recreated tables...\n")

# Main function
def main():
    # Recreate tables
    recreate_tables()
    
    # Insert languages
    insert_languages()
    
    # Insert classes
    insert_classes()
    
    # Insert branches
    insert_branches()

    # Insert data from Excel into MySQL tables
    try:
        insert_excel_data_into_sql_tables()
        print("\nSuccessfully inserted data from Excel into MySQL tables...\n")
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        conn.rollback()
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    main()