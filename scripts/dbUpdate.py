import mysql.connector
from mysql.connector import errorcode
from openpyxl import load_workbook
from dotenv import load_dotenv
import os

# Get the root directory (one level up from the script's directory)
root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

# Load environment variables from the .env.local file in the root directory
load_dotenv(dotenv_path=os.path.join(root_dir, '.env.local'))

# Configs for MySQL connection
config = {
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'host': os.getenv('DB_HOST'),
    'database': os.getenv('DB_NAME'),
    'port': 3306,
    'charset': 'utf8mb4',
    'use_unicode': True
}

# Connect to MySQL server
try:
    conn = mysql.connector.connect(**config)
    read_cursor = conn.cursor(buffered=True)  # Ensure buffered cursor
    write_cursor = conn.cursor()
    print("Successfully connected to MySQL server...\n")
except mysql.connector.Error as err:
    print(f"Error: {err}")
    exit()

# Load Excel spreadsheet
EXCEL_FILENAME = "terms.xlsx"
SHEET_NAME = "Sheet1"
wb = load_workbook(filename=EXCEL_FILENAME, data_only=True)
ws = wb[SHEET_NAME]

# Retrieve branch numbers from the database
read_cursor.execute("SELECT branch_no, map_name FROM branch")
branch_data = read_cursor.fetchall()  # Fetch all results
branch_id_map = {map_name: branch_no for branch_no, map_name in branch_data}

# Retrieve language IDs from the database
read_cursor.execute("SELECT lang_id, name FROM languages")
language_id_map = {name: lang_id for lang_id, name in read_cursor.fetchall()}  # Fetch all results

# Find required column indices
col = 1
term_col = None
branch_col = None
language_cols = []
update_flags = []

while ws.cell(1, col).value is not None:
    header = ws.cell(1, col).value
    if header == "English Terms":
        term_col = col
    elif header == "Field":
        branch_col = col
    else:
        for lang in language_id_map.keys():
            if f"{lang} Definition" in header:
                language_cols.append((col, lang))
            elif f"Has the {lang} translation been edited?" in header:
                update_flags.append((col, lang))
    col += 1

if term_col is None or branch_col is None:
    raise Exception("Missing required columns in Excel sheet...")

# Update translations in the database
for row in range(2, ws.max_row + 1):
    term = ws.cell(row, term_col).value
    branch_name = ws.cell(row, branch_col).value
    branch_no = branch_id_map.get(branch_name)
    
    if not term or not branch_no:
        continue

    # Find term_id
    read_cursor.execute("SELECT term_id FROM terms WHERE term = %s AND branch_no = %s", (term, branch_no))
    term_id = read_cursor.fetchone()  # Always fetch to drain result

    if term_id:
        term_id = term_id[0]
        for (lang_col, lang), (flag_col, flag_lang) in zip(language_cols, update_flags):
            if lang == flag_lang:
                edited_flag = ws.cell(row, flag_col).value
                if edited_flag == "Yes not Uploaded":
                    print(f"Updating translation for term: {term} in {lang}...")
                    new_translation = ws.cell(row, lang_col).value
                    write_cursor.execute(
                        """
                        INSERT INTO translations (term_id, lang_id, definition)
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE definition = %s
                        """,
                        (term_id, language_id_map[lang], new_translation, new_translation)
                    )

conn.commit()
read_cursor.close()
write_cursor.close()
conn.close()
print("Successfully updated translations...")
