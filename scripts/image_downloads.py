import os
import requests
from openpyxl import load_workbook

# Constants
EXCEL_FILENAME = "terms.xlsx"
SHEET_NAME = "Sheet1"
IMAGE_COLUMN = "Mini Link"
TERM_COLUMN = "English Terms"
FIELD_COLUMN = "Field"
SAVE_DIR = "public/terms/mini"

# Load the Excel file
wb = load_workbook(filename=EXCEL_FILENAME, data_only=True)
ws = wb[SHEET_NAME]

# Get header indices
headers = {cell.value: cell.column for cell in ws[1] if cell.value}
print(headers)

if IMAGE_COLUMN not in headers or TERM_COLUMN not in headers or FIELD_COLUMN not in headers:
    raise ValueError("Required columns not found in the Excel sheet.")

image_col_idx = headers[IMAGE_COLUMN]
term_col_idx = headers[TERM_COLUMN]
field_col_idx = headers[FIELD_COLUMN]

# Process each row
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    image_cell = row[image_col_idx - 1]  # Convert Excel 1-based index to 0-based
    term = row[term_col_idx - 1].value
    field = row[field_col_idx - 1].value

    # Extract actual hyperlink (if it exists)
    image_url = None
    if image_cell.hyperlink:
        image_url = image_cell.hyperlink.target

    if not image_url or not term or not field:
        print(f"Skipping row {row_idx} due to missing data.")
        continue

    # Create directory for field
    field_dir = os.path.join(SAVE_DIR, field)
    os.makedirs(field_dir, exist_ok=True)

    # Define image save path
    image_filename = f"{term}.png"
    image_path = os.path.join(field_dir, image_filename)

    # Download image
    try:
        response = requests.get(image_url, timeout=10)
        response.raise_for_status()
        with open(image_path, "wb") as f:
            f.write(response.content)
        print(f"Downloaded: {image_path}")
    except requests.RequestException as e:
        print(f"Failed to download {image_url}: {e}")

print("Image download complete.")
