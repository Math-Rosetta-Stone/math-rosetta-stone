from openpyxl import load_workbook
from enum import Enum
from dataclasses import dataclass

class ApiType(Enum):
    DEEPL = 0
    GOOGLE = 1

@dataclass
class Language:
    language: str
    col_num: int
    api_language_code: str
    api_type: ApiType

# Iniitialize languages to translate to
french = Language("French", 13, "FR", 0)
chinese = Language("Chinese", 11, "ZH", 0)
spanish = Language("Spanish", 15, "ES", 0)
portuguese = Language("Portuguese", 17, "PT-BR", 0)
farsi = Language("Farsi", 19, "fa", 1)
hindi = Language("Hindi", 21, "hi", 1)
marathi = Language("Marathi", 23, "mr", 1)
languages = [french, chinese, spanish, portuguese, hindi, farsi, marathi]


def setup_translation_test(filename, sheetname, en_start_row, en_end_row, ready_to_translate_col):
    """
    Prepares the spreadsheet for testing the translation script.

    Args:
        filename (str): The path to the Excel file.
        sheetname (str): The name of the sheet to modify.
        en_start_row (int): The starting row to process.
        en_end_row (int): The ending row to process.
        ready_to_translate_col (int): The column index for the "Ready to Translate" column.
    """
    # Load the workbook and select the sheet
    wb = load_workbook(filename=filename)
    ws = wb[sheetname]

    # Loop through rows in the specified range
    for row in range(en_start_row, en_end_row + 1):
        # Wipe out translations in the output columns

        for language in languages:
            dst_cell = ws.cell(row=row, column=language.col_num)
            dst_cell.value = None

        # Set "Ready to Translate" column alternately
        ready_cell = ws.cell(row=row, column=ready_to_translate_col)
        if row % 2 == 0:
            ready_cell.value = "Ready for Translation"
        else:
            ready_cell.value = "Not Ready for Translation"

    # Save the workbook
    wb.save(filename)
    print(f"--- Spreadsheet {filename} prepared successfully for testing ---")

# Example usage
setup_translation_test(
    filename="Terms and Definitions.xlsx",
    sheetname="Organized",
    en_start_row=3,
    en_end_row=214,
    ready_to_translate_col=8,
)