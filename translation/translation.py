import argparse
from enum import Enum
from dataclasses import dataclass

import deepl
from google.cloud import translate_v2
from openpyxl import load_workbook

import re


class ApiType(Enum):
    DEEPL = 0
    GOOGLE = 1

@dataclass
class Language:
    language: str
    col_num: int
    api_language_code: str
    api_type: ApiType

start_delim, end_delim = "§§LATEX_START§§", "§§LATEX_END§§"

def extract_text_parts(text):
    """Extracts translatable text portions while keeping LaTeX portions unchanged."""
    parts = re.split(f'({re.escape(start_delim)}.*?{re.escape(end_delim)})', text)
    return parts

def translate_text_parts(parts, translator, language):
    """Translates only the non-LaTeX portions of the text."""
    translated_parts = []
    for part in parts:
        if part.startswith(start_delim) and part.endswith(end_delim):
            translated_parts.append(part)  # Keep LaTeX portions as is
        elif part.strip():  # Only translate if part is not empty
            if language.api_type == ApiType.DEEPL:
                result = translator.translate_text(part, target_lang=language.api_language_code)
                translated_parts.append(result.text)
            elif language.api_type == ApiType.GOOGLE:
                result = translator.translate(part, target_language=language.api_language_code)
                translated_parts.append(result["translatedText"])
            else:
                raise Exception(f"Invalid API type for language {language.language}.")
        else:
            translated_parts.append("")  # Preserve structure without making API calls

    return "".join(translated_parts)

def translate(args):
    # Initialize languages to translate to
    chinese = Language("Chinese", args.zh_out_col, "ZH", ApiType.DEEPL)
    french = Language("French", args.fr_out_col, "FR", ApiType.DEEPL)
    spanish = Language("Spanish", args.es_out_col, "ES", ApiType.DEEPL)
    portuguese = Language("Portuguese", args.pt_out_col, "PT-BR", ApiType.DEEPL)
    farsi = Language("Farsi", args.fa_out_col, "fa", ApiType.GOOGLE)
    hindi = Language("Hindi", args.hi_out_col, "hi", ApiType.GOOGLE)
    marathi = Language("Marathi", args.mr_out_col, "mr", ApiType.GOOGLE)

    languages = [french, chinese, spanish, portuguese, hindi, farsi, marathi]

    # Setup DeepL translator
    auth_key = args.deepl_auth_key
    deepl_translator = deepl.Translator(auth_key)

    # Setup Google Cloud translator
    gc_translator = translate_v2.Client()

    # Load spreadsheet
    wb = load_workbook(filename=args.filename, data_only=True)
    ws = wb[args.sheetname]

    for row in range(args.en_start_row, args.en_end_row + 1):
        # print("ROW:", row)
        src_cell = ws.cell(row=row, column=args.en_col)
        if (src_cell.value is None or
            src_cell.value == "#VALUE!" or
            src_cell.value == 0 or
            src_cell.value == "0"):
            print(f"--- Skipping row {row}, no/invalid definition ---")
            continue

        text_parts = extract_text_parts(src_cell.value)
        # print("TEXT PARTS:", text_parts)
        contains_latex = any(start_delim in part and end_delim in part for part in text_parts)
        is_ready_to_translate = ws.cell(row=row, column=args.ready_to_translate_col)
        if not contains_latex and is_ready_to_translate.value != "Ready for Translation":
            print(f"--- Skipping row {row}, no latex and not ready for translation ---")
            continue

        for language in languages:
            dst_cell = ws.cell(row=row, column=language.col_num)
            is_translation_edited = ws.cell(row=row, column=(language.col_num + 1))

            if contains_latex:
                translated_text = translate_text_parts(text_parts, deepl_translator if language.api_type == ApiType.DEEPL else gc_translator, language)
            else:
                if language.api_type == ApiType.DEEPL:
                    result = deepl_translator.translate_text(src_cell.value, target_lang=language.api_language_code)
                    translated_text = result.text
                elif language.api_type == ApiType.GOOGLE:
                    result = gc_translator.translate(src_cell.value, target_language=language.api_language_code)
                    translated_text = result["translatedText"]
                else:
                    raise Exception(f"Invalid API type for language {language.language}.")

            # Add translated text
            dst_cell.value = translated_text
            # Update is_translation_edited column if specified
            is_translation_edited.value = args.been_edited if args.been_edited else is_translation_edited.value

        # Update "Ready to Translate?" column if specified
        is_ready_to_translate.value = "Translated" if args.update_ready_to_translate else is_ready_to_translate.value

    # Save the updated spreadsheet
    wb.save(args.filename)
    print(f"--- Successfully saved updated spreadsheet to {args.filename} ---")


if __name__ == "__main__":
    # Initialize the argument parser
    parser = argparse.ArgumentParser(description="Translate terms in an Excel sheet.")

    # File arguments
    parser.add_argument(
        "--filename",
        default="Chloe's Copy of the Rosetta Stone Terms.xlsx",
        help="Path to the Excel file."
    )
    parser.add_argument(
        "--sheetname",
        default="Sheet1",
        help="Name of the sheet in the Excel file."
    )

    # Translation input arguments
    parser.add_argument(
        "--ready_to_translate_col",
        default="13",
        type=int,
        help="Column number (1-indexed) to check if a row is ready to translate."
    )
    parser.add_argument(
        "--en_col",
        default="6",
        type=int,
        help="Column number (1-indexed) containing English definitions."
    )
    parser.add_argument(
        "--en_start_row",
        default="2",
        type=int,
        help="Row number (1-indexed) of the first definition to translate."
    )
    parser.add_argument(
        "--en_end_row",
        default="2",
        type=int,
        help="Row number (1-indexed) of the last definition to translate."
    )

    # Translation API arguments - REQUIRED
    parser.add_argument(
        "--deepl_auth_key",
        required=True,
        help="DeepL authentication key."
    )

    # Translation output arguments
    parser.add_argument(
        "--zh_out_col",
        default="14",
        type=int,
        help="Column number (1-indexed) to store Chinese translations."
    )
    parser.add_argument(
        "--fr_out_col",
        default="16",
        type=int,
        help="Column number (1-indexed) to store French translations."
    )
    parser.add_argument(
        "--es_out_col",
        default="18",
        type=int,
        help="Column number (1-indexed) to store Spanish translations."
    )
    parser.add_argument(
        "--pt_out_col",
        default="20",
        type=int,
        help="Column number (1-indexed) to store Portuguese translations."
    )
    parser.add_argument(
        "--fa_out_col",
        default="22",
        type=int,
        help="Column number (1-indexed) to store Farsi translations."
    )
    parser.add_argument(
        "--hi_out_col",
        default="24",
        type=int,
        help="Column number (1-indexed) to store Hindi translations."
    )
    parser.add_argument(
        "--mr_out_col",
        default="26",
        type=int,
        help="Column number (1-indexed) to store Marathi translations."
    )
    parser.add_argument(
        "--been_edited",
        default="",
        help="Value to change the 'Has the translation been edited?' column to. Yes/No. Leave blank to not change."
    )
    parser.add_argument(
        "--update_ready_to_translate",
        action="store_true",
        help="Update 'Ready to translate?' column upon successful translation."
    )

    args = parser.parse_args()
    translate(args)